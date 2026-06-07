import io
import re
import time
from datetime import datetime
from typing import Any

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ALADIN_API_BASE = "http://www.aladin.co.kr/ttb/api/ItemLookUp.aspx"
DEFAULT_SUBJECT_ENDPOINT = "https://apis.data.go.kr/1371029/SubjectInformationService/getSubjectList"

st.set_page_config(page_title="도서목록 주제어 도우미", page_icon="📚", layout="wide")

st.markdown(
    """
    <style>
    .main .block-container {padding-top: 2rem; padding-bottom: 2rem;}
    .small-note {font-size: 0.88rem; color: #666;}
    .ok-box {background:#f3f7ff; border:1px solid #d7e5ff; padding:0.85rem 1rem; border-radius:0.75rem;}
    </style>
    """,
    unsafe_allow_html=True,
)


def normalize_isbn(value: Any) -> str:
    if value is None:
        return ""
    text = re.sub(r"[^0-9Xx]", "", str(value)).strip()
    if len(text) == 13 and text.isdigit():
        return text
    return ""


def read_isbn_file(uploaded_file) -> list[str]:
    name = uploaded_file.name.lower()
    data = uploaded_file.read()
    uploaded_file.seek(0)
    values: list[str] = []
    if name.endswith(".xlsx") or name.endswith(".xls"):
        df = pd.read_excel(io.BytesIO(data), dtype=str)
        if df.empty:
            return []
        candidates = []
        for col in df.columns:
            if "isbn" in str(col).lower() or "바코드" in str(col) or "등록번호" in str(col):
                candidates.append(col)
        if not candidates:
            candidates = [df.columns[0]]
        for col in candidates:
            values.extend(df[col].dropna().astype(str).tolist())
    else:
        text = data.decode("utf-8-sig", errors="ignore")
        values = re.split(r"[\n,\t; ]+", text)
    unique = []
    seen = set()
    for v in values:
        isbn = normalize_isbn(v)
        if isbn and isbn not in seen:
            seen.add(isbn)
            unique.append(isbn)
    return unique


def fetch_aladin_book(isbn: str, ttbkey: str, timeout: int = 12) -> dict[str, Any] | None:
    params = {
        "ttbkey": ttbkey,
        "itemIdType": "ISBN13",
        "ItemId": isbn,
        "output": "js",
        "Version": "20131101",
        "OptResult": "fulldescription,categoryIdList",
    }
    resp = requests.get(ALADIN_API_BASE, params=params, headers={"User-Agent": "Mozilla/5.0"}, timeout=timeout)
    resp.raise_for_status()
    raw = resp.text
    start = raw.find("{")
    end = raw.rfind("}") + 1
    if start == -1 or end <= start:
        return None
    payload = resp.json() if raw.strip().startswith("{") else __import__("json").loads(raw[start:end])
    items = payload.get("item", [])
    return items[0] if items else None


def clean_html(text: Any) -> str:
    if not text:
        return ""
    return BeautifulSoup(str(text), "html.parser").get_text(" ", strip=True)


def parse_aladin_item(item: dict[str, Any], isbn: str) -> dict[str, Any]:
    cat_list = item.get("categoryIdList", [])
    if isinstance(cat_list, list):
        cats = " / ".join(
            str(c.get("categoryName") or c.get("categoryId") or "").strip()
            for c in cat_list
            if str(c.get("categoryName") or c.get("categoryId") or "").strip()
        )
    else:
        cats = str(item.get("categoryName") or "")
    return {
        "ISBN(13자리)": item.get("isbn13") or isbn,
        "서명": item.get("title", ""),
        "저자": item.get("author", ""),
        "출판사": item.get("publisher", ""),
        "출간일": item.get("pubDate") or item.get("pubdate") or "",
        "책소개": clean_html(item.get("fullDescription") or item.get("description") or ""),
        "분야": cats,
    }


def extract_keywords(row: dict[str, Any], max_terms: int = 5) -> list[str]:
    text = " ".join(str(row.get(k, "")) for k in ["서명", "분야", "책소개"])
    text = re.sub(r"[\[\](){}<>『』《》,.:;!?·/|_+=~\"'‘’“”]", " ", text)
    tokens = re.findall(r"[가-힣A-Za-z0-9]{2,}", text)
    stop = {
        "그리고", "그러나", "하지만", "있는", "없는", "위한", "통해", "대한", "에서", "으로", "에게", "우리", "이야기",
        "도서", "책", "개정판", "양장", "세트", "특별판", "완전판", "오늘", "이제", "가장", "모든", "한국", "저자",
    }
    scores: dict[str, int] = {}
    title = str(row.get("서명", ""))
    category = str(row.get("분야", ""))
    for tok in tokens:
        if tok in stop or tok.isdigit() or len(tok) < 2:
            continue
        score = 1
        if tok in title:
            score += 3
        if tok in category:
            score += 2
        if len(tok) >= 4:
            score += 1
        scores[tok] = scores.get(tok, 0) + score
    return [k for k, _ in sorted(scores.items(), key=lambda x: (-x[1], x[0]))[:max_terms]]


def _tag_local_name(tag: Any) -> str:
    name = getattr(tag, "name", tag)
    name = str(name or "")
    for sep in ["#", "/", ":", "."]:
        if sep in name:
            name = name.split(sep)[-1]
    return name.strip()


BAD_VALUE_PATTERNS = [
    "markusinginnlk", "http", "https", "rdf", "skos", "owl", "xmlns", "subjectinformationservice",
    "apis.data.go.kr", "www.nl.go.kr", "nlsht", "resource", "description",
]


def _clean_subject_value(value: Any) -> str:
    if value is None:
        return ""
    text = clean_html(str(value))
    text = re.sub(r"https?://\S+", " ", text, flags=re.I)
    text = re.sub(r"[{}\[\]'\"<>]", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" |,;/")
    if not text:
        return ""
    low = text.lower()
    if any(bad in low for bad in BAD_VALUE_PATTERNS) and not re.search(r"[가-힣]", text):
        return ""
    parts = [p.strip() for p in re.split(r"\s*[|;/,]\s*", text) if p.strip()]
    hangul_parts = [p for p in parts if re.search(r"[가-힣]", p) and not any(bad in p.lower() for bad in BAD_VALUE_PATTERNS)]
    if hangul_parts:
        return " | ".join(dict.fromkeys(hangul_parts[:8]))
    if any(bad in low for bad in BAD_VALUE_PATTERNS):
        return ""
    return text


def _join_clean(values: list[Any]) -> str:
    cleaned = []
    for v in values:
        c = _clean_subject_value(v)
        if c:
            for part in c.split(" | "):
                if part and part not in cleaned:
                    cleaned.append(part)
    return " | ".join(cleaned[:12])


def xml_to_dict_list(text: str) -> list[dict[str, Any]]:
    soup = BeautifulSoup(text, "xml")
    candidates = soup.find_all(["item", "row", "data", "subject", "result", "Description"])
    if not candidates:
        candidates = soup.find_all(lambda tag: tag.find_all(recursive=False))[:50]

    rows: list[dict[str, Any]] = []
    for node in candidates:
        row: dict[str, Any] = {}
        for child in node.find_all(recursive=False):
            key = _tag_local_name(child)
            val = child.get_text(" ", strip=True)
            if not val:
                for attr_name, attr_val in child.attrs.items():
                    if str(attr_name).lower().endswith(("resource", "about", "id")):
                        val = attr_val
                        break
            if key:
                if key in row and val:
                    row[key] = f"{row[key]} | {val}"
                else:
                    row[key] = val
        if row:
            rows.append(row)
    if not rows:
        rows.append({"raw": soup.get_text(" ", strip=True)[:1000]})
    return rows


def find_records(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [x for x in payload if isinstance(x, dict)]
    if not isinstance(payload, dict):
        return []
    keys = ["items", "item", "data", "result", "results", "list", "row", "body", "subjects", "subject", "@graph", "graph"]
    queue = [payload]
    seen = set()
    while queue:
        cur = queue.pop(0)
        oid = id(cur)
        if oid in seen:
            continue
        seen.add(oid)
        if isinstance(cur, dict):
            for key in keys:
                val = cur.get(key)
                if isinstance(val, list):
                    dicts = [x for x in val if isinstance(x, dict)]
                    if dicts:
                        return dicts
                elif isinstance(val, dict):
                    queue.append(val)
            for val in cur.values():
                if isinstance(val, dict):
                    queue.append(val)
                elif isinstance(val, list):
                    dicts = [x for x in val if isinstance(x, dict)]
                    if dicts:
                        return dicts
    return [payload]


def flatten_record(d: dict[str, Any], prefix: str = "") -> dict[str, str]:
    out: dict[str, str] = {}
    for k, v in d.items():
        key_name = _tag_local_name(k)
        key = f"{prefix}.{key_name}" if prefix else key_name
        if isinstance(v, dict):
            for value_key in ["@value", "value", "label", "Label", "name"]:
                if value_key in v and not isinstance(v.get(value_key), (dict, list)):
                    out[key] = str(v.get(value_key) or "")
                    break
            else:
                out.update(flatten_record(v, key))
        elif isinstance(v, list):
            simple = []
            for item in v:
                if isinstance(item, dict):
                    picked = ""
                    for value_key in ["@value", "value", "label", "prefLabel", "altLabel", "name", "@id", "id"]:
                        if value_key in item and not isinstance(item.get(value_key), (dict, list)):
                            picked = str(item.get(value_key) or "")
                            break
                    if not picked:
                        inner = flatten_record(item)
                        picked = " | ".join(inner.values())
                    simple.append(picked)
                else:
                    simple.append(str(item))
            out[key] = " | ".join(simple)
        else:
            out[key] = "" if v is None else str(v)
    return out


def pick_first(record: dict[str, str], names: list[str]) -> str:
    candidates: list[str] = []
    lowered = {k.lower(): v for k, v in record.items()}
    for name in names:
        target = _tag_local_name(name).lower()
        for k, v in record.items():
            local = _tag_local_name(k).lower()
            if (local == target or local.endswith(target)) and v:
                candidates.append(v)
        for k, v in lowered.items():
            local = _tag_local_name(k).lower()
            if (local == target or local.endswith(target)) and v:
                candidates.append(v)
    cleaned = [_clean_subject_value(v) for v in candidates]
    cleaned = [v for v in cleaned if v]
    hangul = [v for v in cleaned if re.search(r"[가-힣]", v)]
    if hangul:
        return hangul[0]
    return cleaned[0] if cleaned else ""


def normalize_subject_records(keyword: str, records: list[dict[str, Any]]) -> list[dict[str, str]]:
    rows = []
    for rec in records:
        flat = flatten_record(rec)
        preferred = pick_first(flat, ["label", "prefLabel", "preferredTerm", "subject", "subjectName", "name", "title", "heading", "term", "주제명", "우선어"])
        uf = pick_first(flat, ["altLabel", "useFor", "uf", "비우선어"])
        bt = pick_first(flat, ["broader", "broaderLabel", "bt", "상위어"])
        nt = pick_first(flat, ["narrower", "narrowerLabel", "nt", "하위어"])
        rt = pick_first(flat, ["related", "relatedLabel", "rt", "관련어"])
        uri = pick_first(flat, ["subjectId", "id", "identifier", "uri", "about"])

        if not preferred:
            raw_text = " ".join(str(v) for v in flat.values())
            terms = re.findall(r"[가-힣][가-힣A-Za-z0-9·\- ]{1,30}", raw_text)
            terms = [t.strip() for t in terms if t.strip() and len(t.strip()) >= 2]
            if terms:
                preferred = terms[0]

        summary = _join_clean([preferred, uf, bt, nt, rt])
        if not any([preferred, uf, bt, nt, rt, uri]):
            summary = _join_clean(list(flat.values())) or str(flat)[:500]
        rows.append({
            "조회키워드": keyword,
            "표준주제어_후보": preferred,
            "비우선어_UF": uf,
            "상위어_BT": bt,
            "하위어_NT": nt,
            "관련어_RT": rt,
            "식별자_URI": uri,
            "원문요약": summary,
        })
    return [r for r in rows if any(r.get(k) for k in ["표준주제어_후보", "비우선어_UF", "상위어_BT", "하위어_NT", "관련어_RT", "식별자_URI", "원문요약"])]


def build_subject_params(service_key: str, keyword: str, param_name: str, page_param: str, size_param: str, rows: int, extra_params: str, response_param: str, response_value: str) -> dict[str, Any]:
    params: dict[str, Any] = {"serviceKey": service_key}
    if param_name:
        params[param_name] = keyword
    if page_param:
        params[page_param] = 1
    if size_param:
        params[size_param] = rows
    if response_param and response_value:
        params[response_param] = response_value
    for part in extra_params.split("&"):
        if "=" in part:
            k, v = part.split("=", 1)
            if k.strip():
                params[k.strip()] = v.strip()
    return params


def preview_subject_url(endpoint: str, params: dict[str, Any], key_mode: str) -> str:
    safe = dict(params)
    safe["serviceKey"] = "***API_KEY***"
    if key_mode == "encoded":
        sep = "&" if "?" in endpoint else "?"
        url = f"{endpoint.strip()}{sep}serviceKey=***API_KEY***"
        return requests.Request("GET", url, params={k: v for k, v in safe.items() if k != "serviceKey"}).prepare().url or url
    return requests.Request("GET", endpoint.strip(), params=safe).prepare().url or endpoint.strip()


def call_subject_api(endpoint: str, service_key: str, keyword: str, param_name: str, page_param: str, size_param: str, rows: int, extra_params: str, response_param: str = "returnType", response_value: str = "json", key_mode: str = "decoding") -> list[dict[str, str]]:
    if not endpoint.strip():
        return []
    params = build_subject_params(service_key, keyword, param_name, page_param, size_param, rows, extra_params, response_param, response_value)
    if key_mode == "encoded":
        sep = "&" if "?" in endpoint else "?"
        url = f"{endpoint.strip()}{sep}serviceKey={service_key}"
        resp = requests.get(url, params={k: v for k, v in params.items() if k != "serviceKey"}, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
    else:
        resp = requests.get(endpoint.strip(), params=params, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
    if resp.status_code >= 400:
        req_url = resp.url.replace(service_key, "***API_KEY***")
        body = resp.text[:1000].replace(service_key, "***API_KEY***")
        raise RuntimeError(f"HTTP {resp.status_code}: {resp.reason}\n요청URL: {req_url}\n응답본문: {body}")
    content_type = resp.headers.get("content-type", "").lower()
    if "json" in content_type or resp.text.strip().startswith(("{", "[")):
        payload = resp.json()
        records = find_records(payload)
    else:
        records = xml_to_dict_list(resp.text)
    return normalize_subject_records(keyword, records)[:rows]


def make_excel(book_rows: list[dict[str, Any]], fail_rows: list[dict[str, Any]], filename: str = "도서목록_주제어_결과.xlsx") -> bytes:
    output = io.BytesIO()
    df_books = pd.DataFrame(book_rows)
    df_fail = pd.DataFrame(fail_rows)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_books.to_excel(writer, index=False, sheet_name="도서정보_주제어")
        if not df_fail.empty:
            df_fail.to_excel(writer, index=False, sheet_name="조회실패")
        wb = writer.book
        for ws in wb.worksheets:
            header_fill = PatternFill("solid", fgColor="D9EAF7")
            thin = Side(style="thin", color="D9D9D9")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for cell in ws[1]:
                cell.font = Font(name="맑은 고딕", size=10, bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(name="맑은 고딕", size=10)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = border
            for idx, col in enumerate(ws.columns, start=1):
                max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col[:60])
                width = min(max(max_len + 2, 10), 42)
                if ws.cell(row=1, column=idx).value in ["책소개", "원문요약"]:
                    width = 55
                ws.column_dimensions[get_column_letter(idx)].width = width
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
    return output.getvalue()


st.title("📚 도서목록 주제어 도우미")
st.caption("ISBN 엑셀/텍스트 업로드 → 알라딘 도서정보 조회 → 국립중앙도서관 주제정보 후보 붙이기 → 엑셀 다운로드")

with st.sidebar:
    st.header("1. API 키")
    aladin_key = st.text_input("알라딘 TTBKey", type="password", help="입력값은 코드에 저장되지 않습니다.")
    public_key = st.text_input("공공데이터포털 인증키", type="password", help="가이드상 일반 인증키(Decoding)를 권장합니다. 안 되면 Encoding 키로도 테스트할 수 있습니다.")
    key_mode = st.radio("공공데이터포털 키 형식", ["decoding", "encoded"], format_func=lambda x: "일반 인증키(Decoding)" if x == "decoding" else "인코딩 인증키(Encoding)", horizontal=True)
    st.divider()
    st.header("2. 주제정보 API 설정")
    subject_endpoint = st.text_input("주제정보 API 호출 URL", value=DEFAULT_SUBJECT_ENDPOINT, help="Swagger의 Request URL에서 ? 앞부분까지 붙여넣으세요. 예: https://apis.data.go.kr/.../...")
    keyword_param = st.text_input("검색어 요청변수명", value="label", help="이 API는 보통 label을 씁니다.")
    page_param = st.text_input("페이지 요청변수명", value="pageNo")
    size_param = st.text_input("건수 요청변수명", value="numOfRows")
    response_param = st.text_input("응답형식 요청변수명", value="type", help="이 API는 보통 type=json 또는 type=xml을 씁니다. 없으면 비우세요.")
    response_value = st.text_input("응답형식 값", value="json", help="json 또는 xml. Swagger에서 쓰는 값을 넣으세요.")
    extra_params = st.text_input("추가 요청변수", value="", help="예: type=subject&format=json 처럼 &로 연결. Swagger 필수값이 더 있으면 여기에 넣으세요.")
    subject_rows = st.slider("키워드별 주제어 후보 수", 1, 10, 3)
    max_keywords = st.slider("도서별 자동 추출 키워드 수", 1, 8, 4)
    delay_sec = st.slider("API 요청 간격(초)", 0.0, 2.0, 0.3, 0.1)

uploaded = st.file_uploader("ISBN 파일 업로드", type=["xlsx", "xls", "txt", "csv"])

col_a, col_b, col_c = st.columns([1, 1, 2])
with col_a:
    run_aladin = st.checkbox("알라딘 도서정보 조회", value=True)
with col_b:
    run_subject = st.checkbox("주제정보 API 조회", value=True)
with col_c:
    st.markdown('<div class="small-note">엑셀은 ISBN 컬럼이 있으면 우선 사용하고, 없으면 첫 번째 컬럼을 읽습니다.</div>', unsafe_allow_html=True)

if uploaded:
    isbns = read_isbn_file(uploaded)
    st.success(f"유효한 13자리 ISBN {len(isbns)}개를 찾았습니다.")
    with st.expander("읽어들인 ISBN 미리보기"):
        st.write(isbns[:50])
else:
    isbns = []

st.subheader("주제정보 API 단독 테스트")
test_keyword = st.text_input("테스트 키워드", value="환경")
if st.button("주제어 테스트 조회"):
    if not public_key or not subject_endpoint or subject_endpoint == DEFAULT_SUBJECT_ENDPOINT:
        st.warning("공공데이터포털 인증키와 실제 주제정보 API 호출 URL을 입력하세요.")
    else:
        try:
            preview_params = build_subject_params(public_key, test_keyword, keyword_param, page_param, size_param, subject_rows, extra_params, response_param, response_value)
            st.caption("실제 호출 URL 미리보기(API키는 가림)")
            st.code(preview_subject_url(subject_endpoint, preview_params, key_mode), language="text")
            rows = call_subject_api(subject_endpoint, public_key, test_keyword, keyword_param, page_param, size_param, subject_rows, extra_params, response_param, response_value, key_mode)
            st.dataframe(pd.DataFrame(rows), use_container_width=True)
        except Exception as e:
            st.error(f"조회 오류: {e}")
            st.info("대부분 호출 URL, 검색어 요청변수명, 응답형식 변수명(returnType/_type 등), 인증키 형식(Decoding/Encoding) 불일치에서 발생합니다. 위에 표시된 요청URL을 Swagger의 Request URL과 비교해 주세요.")

st.divider()

if st.button("도서목록 주제어 생성", type="primary", disabled=not bool(isbns)):
    if run_aladin and not aladin_key:
        st.error("알라딘 TTBKey를 입력하세요.")
        st.stop()
    if run_subject and (not public_key or not subject_endpoint or subject_endpoint == DEFAULT_SUBJECT_ENDPOINT):
        st.error("공공데이터포털 인증키와 실제 주제정보 API 호출 URL을 입력하세요.")
        st.stop()

    progress = st.progress(0)
    status = st.empty()
    results: list[dict[str, Any]] = []
    fails: list[dict[str, Any]] = []

    for i, isbn in enumerate(isbns, start=1):
        status.write(f"[{i}/{len(isbns)}] {isbn} 처리 중")
        try:
            if run_aladin:
                item = fetch_aladin_book(isbn, aladin_key)
                if not item:
                    fails.append({"ISBN(13자리)": isbn, "단계": "알라딘", "사유": "결과 없음"})
                    progress.progress(i / len(isbns))
                    time.sleep(delay_sec)
                    continue
                row = parse_aladin_item(item, isbn)
            else:
                row = {"ISBN(13자리)": isbn, "서명": "", "저자": "", "출판사": "", "출간일": "", "책소개": "", "분야": ""}

            keywords = extract_keywords(row, max_terms=max_keywords)
            row["자동추출키워드"] = ", ".join(keywords)

            subject_candidates: list[dict[str, str]] = []
            if run_subject:
                for kw in keywords:
                    try:
                        subject_candidates.extend(call_subject_api(subject_endpoint, public_key, kw, keyword_param, page_param, size_param, subject_rows, extra_params, response_param, response_value, key_mode))
                    except Exception as e:
                        fails.append({"ISBN(13자리)": isbn, "단계": f"주제정보:{kw}", "사유": str(e)[:400]})
                    time.sleep(delay_sec)

            row["표준주제어_후보"] = " | ".join(dict.fromkeys([x.get("표준주제어_후보", "") for x in subject_candidates if x.get("표준주제어_후보")]))
            row["상위어_BT"] = " | ".join(dict.fromkeys([x.get("상위어_BT", "") for x in subject_candidates if x.get("상위어_BT")]))
            row["하위어_NT"] = " | ".join(dict.fromkeys([x.get("하위어_NT", "") for x in subject_candidates if x.get("하위어_NT")]))
            row["관련어_RT"] = " | ".join(dict.fromkeys([x.get("관련어_RT", "") for x in subject_candidates if x.get("관련어_RT")]))
            row["비우선어_UF"] = " | ".join(dict.fromkeys([x.get("비우선어_UF", "") for x in subject_candidates if x.get("비우선어_UF")]))
            row["주제정보_원문요약"] = " || ".join([x.get("원문요약", "") for x in subject_candidates[:8] if x.get("원문요약")])
            results.append(row)
        except Exception as e:
            fails.append({"ISBN(13자리)": isbn, "단계": "전체", "사유": str(e)[:400]})
        progress.progress(i / len(isbns))
        time.sleep(delay_sec)

    df = pd.DataFrame(results)
    st.subheader("결과 미리보기")
    st.dataframe(df, use_container_width=True)
    if fails:
        st.warning(f"조회 실패/부분 오류 {len(fails)}건이 있습니다. 엑셀의 조회실패 시트에서 확인하세요.")
        with st.expander("실패 목록 보기"):
            st.dataframe(pd.DataFrame(fails), use_container_width=True)

    excel_bytes = make_excel(results, fails)
    out_name = f"도서목록_주제어_결과_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    st.download_button("엑셀 다운로드", data=excel_bytes, file_name=out_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
