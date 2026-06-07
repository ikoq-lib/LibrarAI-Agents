"""
알라딘 Open API - ISBN 일괄 조회 스크립트
=========================================
사용법:
  python aladin_book_lookup.py <ISBN파일.txt> <TTBKey>

예시:
  python aladin_book_lookup.py isbns.txt ttbmykey1234567

ISBN 파일 형식:
  한 줄에 13자리 ISBN 하나씩
  예)
    9791199239098
    9791199525900
    9791169095433

결과:
  알라딘_도서정보_YYYY-MM-DD.xlsx 로 저장됨
"""

import sys
import time
import re
import json
import urllib.request
import urllib.parse
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("[오류] openpyxl 이 없습니다. 아래 명령어로 설치 후 다시 실행하세요:")
    print("  pip install openpyxl")
    sys.exit(1)


# ── 설정 ──────────────────────────────────────────────────────────────────────
DELAY_SEC   = 0.3   # 요청 간 딜레이 (초)
API_BASE    = "http://www.aladin.co.kr/ttb/api/ItemLookUp.aspx"
OUTPUT_FILE = f"알라딘_도서정보_{date.today()}.xlsx"
# ──────────────────────────────────────────────────────────────────────────────


def fetch_book(isbn: str, ttbkey: str) -> dict | None:
    params = urllib.parse.urlencode({
        "ttbkey":     ttbkey,
        "itemIdType": "ISBN13",
        "ItemId":     isbn,
        "output":     "js",
        "Version":    "20131101",
        "OptResult":  "fulldescription,categoryIdList",
    })
    url = f"{API_BASE}?{params}"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        raw = resp.read().decode("utf-8")

    # JSONP 형태 제거 → 순수 JSON 추출
    start = raw.find("{")
    end   = raw.rfind("}") + 1
    if start == -1:
        return None
    data = json.loads(raw[start:end])
    items = data.get("item", [])
    return items[0] if items else None


def parse_item(item: dict) -> dict:
    # 분야 파싱
    cat_list = item.get("categoryIdList", [])
    if isinstance(cat_list, list):
        cats = " / ".join(
            c.get("categoryName", str(c.get("categoryId", "")))
            for c in cat_list
        )
    else:
        cats = ""

    # 책소개 HTML 태그 제거
    desc = item.get("fullDescription") or item.get("description") or ""
    desc = re.sub(r"<[^>]+>", "", desc).strip()

    return {
        "isbn13":           item.get("isbn13", ""),
        "title":            item.get("title", ""),
        "author":           item.get("author", ""),
        "publisher":        item.get("publisher", ""),
        "pubdate":          item.get("pubdate", ""),
        "fullDescription":  desc,
        "category":         cats,
    }


def save_excel(rows: list[dict], fail_isbns: list[str]):
    wb = openpyxl.Workbook()

    # ── 시트1: 도서정보 ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "도서정보"

    headers = ["ISBN(13자리)", "서명", "저자", "출판사", "출간일", "책소개", "분야"]
    col_widths = [18, 36, 22, 18, 12, 70, 32]

    header_font  = Font(name="맑은 고딕", bold=True, size=10)
    header_fill  = PatternFill("solid", fgColor="D9E1F2")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font    = Font(name="맑은 고딕", size=10)
    body_align   = Alignment(vertical="top", wrap_text=True)
    thin         = Side(style="thin", color="BFBFBF")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 22
    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    for ri, row in enumerate(rows, start=2):
        values = [
            row["isbn13"], row["title"], row["author"],
            row["publisher"], row["pubdate"],
            row["fullDescription"], row["category"],
        ]
        ws.row_dimensions[ri].height = 60
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = body_font
            cell.alignment = body_align
            cell.border    = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── 시트2: 조회실패 ────────────────────────────────────────────────────────
    if fail_isbns:
        ws2 = wb.create_sheet("조회실패")
        ws2.cell(row=1, column=1, value="ISBN(13자리)").font = header_font
        ws2.cell(row=1, column=1).fill = PatternFill("solid", fgColor="FCE4D6")
        ws2.column_dimensions["A"].width = 18
        for ri, isbn in enumerate(fail_isbns, start=2):
            ws2.cell(row=ri, column=1, value=isbn).font = body_font

    wb.save(OUTPUT_FILE)


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    isbn_file = sys.argv[1]
    ttbkey    = sys.argv[2]

    # ISBN 읽기
    with open(isbn_file, encoding="utf-8") as f:
        raw_lines = f.read().splitlines()
    isbns = list(dict.fromkeys(
        l.strip() for l in raw_lines if re.fullmatch(r"\d{13}", l.strip())
    ))

    if not isbns:
        print("[오류] 유효한 13자리 ISBN을 찾을 수 없습니다.")
        sys.exit(1)

    print(f"총 {len(isbns)}개 ISBN 조회 시작\n")

    success_rows = []
    fail_isbns   = []

    for i, isbn in enumerate(isbns, start=1):
        try:
            item = fetch_book(isbn, ttbkey)
            if item:
                success_rows.append(parse_item(item))
                print(f"  ✓ [{i}/{len(isbns)}] {item.get('title', isbn)}")
            else:
                fail_isbns.append(isbn)
                print(f"  - [{i}/{len(isbns)}] {isbn}  (결과 없음)")
        except Exception as e:
            fail_isbns.append(isbn)
            print(f"  ✗ [{i}/{len(isbns)}] {isbn}  오류: {e}")

        time.sleep(DELAY_SEC)

    save_excel(success_rows, fail_isbns)

    print(f"\n완료: 성공 {len(success_rows)}개 / 실패 {len(fail_isbns)}개")
    print(f"저장 위치: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
