# -*- coding: utf-8 -*-
"""정기구입 심의대상 목록(ATT-009 서식) 생성.

acquisition_shortlist_kdc() 결과 JSON 을 받아 실물 서식과 같은 구조의 xlsx 를 만든다.
  총괄표 / 일반 / 어린이  = ATT-009 원본 3시트 구조 그대로
  선정근거                = 신규. 6개 축 점수와 근거 지표(회전율·결핍·추천수)

원본 서식: templates/official_docs/ATT-009_자료개발_정기구입심의대상목록_서식.xlsx

사용:  python scripts/build_acquisition_review_list.py <입력.json> <출력.xlsx> [제목]
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="999999")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="DDEBF7")
TITLE_FONT = Font(size=14, bold=True)
HEAD_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# KDC 대분류 표기 순서 (ATT-009 총괄표와 동일)
KDC_ORDER = [
    ("0", "총류"), ("1", "철학"), ("2", "종교"), ("3", "사회과학"), ("4", "자연과학"),
    ("5", "기술과학"), ("6", "예술"), ("7", "언어"), ("8", "문학"), ("9", "역사"),
]


def pub_year(row):
    d = row.get("publish_predate") or ""
    return d[:4] if len(d) >= 4 else ""


# SEOJI AUTHOR 는 "저자 : 김애란;삽화가(그림작가) : 유영근;" 형태의 역할 구조 문자열이다.
# 심의목록에는 실물 서식처럼 저자명만 넣는다(1명 그대로, 2명 나열, 3명 이상 "외").
AUTHOR_ROLES = ("저자", "지은이", "글", "원작", "편저", "엮은이")


def clean_author(raw):
    if not raw:
        return ""
    names, fallback = [], []
    for part in str(raw).split(";"):
        part = part.strip()
        if not part:
            continue
        if ":" in part:
            role, _, name = part.partition(":")
            role, name = role.strip(), name.strip()
        else:
            role, name = "", part
        if not name:
            continue
        fallback.append(name)
        if any(k in role for k in AUTHOR_ROLES):
            names.append(name)
    picked = names or fallback[:1]
    if not picked:
        return str(raw).strip()
    if len(picked) == 1:
        return picked[0]
    if len(picked) == 2:
        return "%s, %s" % (picked[0], picked[1])
    return "%s 외" % picked[0]


def price(row):
    return int(row.get("pre_price") or 0)


def style_header(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row_idx, c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BOX


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_general(wb, rows, title):
    """일반 시트 - ATT-009 열 구조(순/주제/서명/저자/출판사/출판년/ISBN/책수/가격/비고) + 선정점수."""
    ws = wb.create_sheet("일반")
    ws["A1"] = title + "(일반)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws["A1"].alignment = CENTER

    head = ["순", "주제", "서명", "저자", "출판사", "출판년", "ISBN", "책수", "가격", "비고", "선정점수"]
    ws.append([])
    ws.append(head)
    style_header(ws, 2, len(head))

    for i, r in enumerate(rows, start=1):
        ws.append([
            i, r.get("주제") or "미상", r.get("title"), clean_author(r.get("author")), r.get("publisher"),
            pub_year(r), r.get("ea_isbn"), 1, price(r),
            "청소년" if r.get("대상") == "청소년" else "",
            round(float(r.get("score_total") or 0), 1),
        ])
    last = ws.max_row
    ws.append(["", "계", "", "", "", "", "", "=SUM(H3:H%d)" % last, "=SUM(I3:I%d)" % last, "", ""])
    for c in (2, 8, 9):
        ws.cell(ws.max_row, c).font = HEAD_FONT
    set_widths(ws, [5, 10, 46, 20, 18, 8, 16, 6, 10, 8, 9])
    ws.freeze_panes = "A3"


def sheet_child(wb, rows, title):
    """어린이 시트 - ATT-009 열 구조(순/분야/서명/저자/출판사/출판년/단가/권수/금액/비고) + 주제·선정점수."""
    ws = wb.create_sheet("어린이")
    ws["A1"] = title + "(어린이)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws["A1"].alignment = CENTER

    head = ["순", "분야", "서명", "저자", "출판사", "출판년", "단가", "권수", "금액", "비고", "주제", "선정점수"]
    ws.append([])
    ws.append(head)
    style_header(ws, 2, len(head))

    for i, r in enumerate(rows, start=1):
        rw = i + 2
        ws.append([
            i, "어린이", r.get("title"), clean_author(r.get("author")), r.get("publisher"),
            pub_year(r), price(r), 1, "=G%d*H%d" % (rw, rw), "",
            r.get("주제") or "미상", round(float(r.get("score_total") or 0), 1),
        ])
    last = ws.max_row
    ws.append(["", "계", "", "", "", "", "", "=SUM(H3:H%d)" % last, "=SUM(I3:I%d)" % last, "", "", ""])
    for c in (2, 8, 9):
        ws.cell(ws.max_row, c).font = HEAD_FONT
    set_widths(ws, [5, 9, 46, 20, 18, 8, 10, 6, 12, 8, 10, 9])
    ws.freeze_panes = "A3"


def sheet_basis(wb, rows):
    """선정근거 - 6개 축 점수와 근거 지표."""
    ws = wb.create_sheet("선정근거")
    head = ["순", "서명", "주제", "대상", "총점", "출판사(20)", "저자(20)", "발행일(20)",
            "인기도(15)", "균형(15)", "가격(10)",
            "출판사소장", "출판사회전율", "저자소장", "저자회전율", "KDC결핍", "추천출처수", "미상축",
            "저자(원문)"]
    ws.append(head)
    style_header(ws, 1, len(head))

    def num(r, k, nd=1):
        v = r.get(k)
        return round(float(v), nd) if v is not None else None

    for i, r in enumerate(rows, start=1):
        unknown = r.get("unknown_axes")
        if isinstance(unknown, list):
            unknown = ", ".join(unknown)
        ws.append([
            i, r.get("title"), r.get("주제") or "미상", r.get("대상"),
            num(r, "score_total"), num(r, "score_publisher"), num(r, "score_author"),
            num(r, "score_pubdate"), num(r, "score_popularity"), num(r, "score_kdc"),
            num(r, "score_price"),
            r.get("pub_holdings"), num(r, "pub_turnover", 3),
            r.get("aut_holdings"), num(r, "aut_turnover", 3),
            num(r, "kdc_deficit", 4), r.get("recommend_count"), unknown, r.get("author"),
        ])
    set_widths(ws, [5, 42, 10, 8, 8, 11, 10, 11, 11, 9, 9, 11, 12, 10, 11, 10, 11, 10, 40])
    ws.freeze_panes = "C2"


def sheet_summary(wb, general, child, title, note_lines):
    """총괄표 - ATT-009 원본 구조."""
    ws = wb.create_sheet("총괄표", 0)
    total = general + child
    amount = sum(price(r) for r in total)

    ws["A1"] = title + " 현황"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "1. 소요예산: 금{:,}원".format(amount)
    ws["A3"] = "2. 수량: 총 {}권".format(len(total))
    ws["A4"] = "3. 대상별 자료 현황"

    # 대상별 - 유아/어린이/청소년/일반
    groups = {"유아": [], "어린이": [], "청소년": [], "일반": []}
    for r in total:
        t = r.get("대상")
        if t == "아동":
            groups["어린이"].append(r)
        elif t == "청소년":
            groups["청소년"].append(r)
        else:
            groups["일반"].append(r)

    ws["A5"] = "구     분"
    for i, g in enumerate(["유아", "어린이", "청소년", "일반"], start=2):
        ws.cell(5, i, g)
    ws.cell(5, 6, "합   계")
    ws["A6"] = "수량(권)"
    ws["A7"] = "금액(원)"
    for i, g in enumerate(["유아", "어린이", "청소년", "일반"], start=2):
        ws.cell(6, i, len(groups[g]))
        ws.cell(7, i, sum(price(r) for r in groups[g]))
    ws.cell(6, 6, "=SUM(B6:E6)")
    ws.cell(7, 6, "=SUM(B7:E7)")
    for rr in (5, 6, 7):
        for cc in range(1, 7):
            ws.cell(rr, cc).border = BOX
            ws.cell(rr, cc).alignment = CENTER
    style_header(ws, 5, 6)

    ws["A8"] = "4. 류별 자료 현황"
    ws["A9"] = "대상"
    for i, h in enumerate(["구분", "종수", "권수", "금액(원)", "비율(%)", "비고"], start=2):
        ws.cell(9, i, h)
    style_header(ws, 9, 7)

    row = 10
    ws.cell(row, 1, "일반도서\n(청소년 포함)")
    for code, label in KDC_ORDER:
        sub = [r for r in general if (r.get("kdc_major") or "") == code]
        ws.cell(row, 2, label)
        ws.cell(row, 3, len(sub))
        ws.cell(row, 4, len(sub))
        ws.cell(row, 5, sum(price(r) for r in sub))
        ws.cell(row, 6, "=IF($D$23=0,0,D%d/$D$23*100)" % row)
        row += 1
    ws.cell(row, 2, "소계")
    ws.cell(row, 3, "=SUM(C10:C19)")
    ws.cell(row, 4, "=SUM(D10:D19)")
    ws.cell(row, 5, "=SUM(E10:E19)")
    ws.cell(row, 6, "=IF($D$23=0,0,D%d/$D$23*100)" % row)
    row += 1  # 21

    ws.cell(row, 1, "어린이도서\n(유아 포함)")
    ws.cell(row, 2, "어린이")
    ws.cell(row, 3, len(child))
    ws.cell(row, 4, len(child))
    ws.cell(row, 5, sum(price(r) for r in child))
    ws.cell(row, 6, "=IF($D$23=0,0,D%d/$D$23*100)" % row)
    row += 1  # 22
    ws.cell(row, 2, "소계")
    ws.cell(row, 3, "=C21")
    ws.cell(row, 4, "=D21")
    ws.cell(row, 5, "=E21")
    ws.cell(row, 6, "=IF($D$23=0,0,D%d/$D$23*100)" % row)
    row += 1  # 23
    ws.cell(row, 1, "합계")
    ws.cell(row, 3, "=C20+C22")
    ws.cell(row, 4, "=D20+D22")
    ws.cell(row, 5, "=E20+E22")
    ws.cell(row, 6, "=F20+F22")

    for rr in range(9, 24):
        for cc in range(1, 8):
            ws.cell(rr, cc).border = BOX
            ws.cell(rr, cc).alignment = CENTER
    ws.merge_cells(start_row=10, start_column=1, end_row=20, end_column=1)
    ws.merge_cells(start_row=21, start_column=1, end_row=22, end_column=1)

    r = 25
    ws.cell(r, 1, "※ 산출 근거 및 한계")
    ws.cell(r, 1).font = HEAD_FONT
    for line in note_lines:
        r += 1
        ws.cell(r, 1, line)
    set_widths(ws, [18, 14, 10, 10, 14, 12, 24])


NOTES = [
    "· 후보 풀: 국립중앙도서관 SEOJI 종이책 신간 수집분",
    "· 선정: acquisition_shortlist_kdc() - KDC 목표비율(ATT-006 산식 D)을 권수 쿼터로 강제하고 류 안에서 점수순",
    "· 점수 6개 축: 출판사 20 / 저자 20 / 발행일 20 / 인기도 15 / KDC균형 15 / 가격 10",
    "· 회전율은 우리 관 실제 대출(2024-01~2026-07, 131,092건)과 소장 73,390권 기준",
    "· 한계 1: SEOJI 부가기호로는 유아와 어린이를 구분할 수 없어 유아 수량을 0으로 두고 어린이에 합산했다.",
    "· 한계 2: 금액은 SEOJI 정가이며 실제 구입가(할인율)는 반영하지 않았다.",
    "· 한계 3: 복본 판정은 ISBN 완전일치만 반영했다. 서명+저자 유사도 판정(B-03)은 미적용.",
    "· 이 목록은 심의 대상 후보이며 최종 선정은 자료심의위원회와 사서가 결정한다.",
]


def main():
    src, dst = sys.argv[1], sys.argv[2]
    title = sys.argv[3] if len(sys.argv) > 3 else "정기구입 자료 심의대상 목록"
    with open(src, encoding="utf-8") as f:
        rows = json.load(f)
    rows.sort(key=lambda r: ((r.get("kdc_major") or "9"), -float(r.get("score_total") or 0)))

    child = [r for r in rows if r.get("대상구분") == "child"]
    general = [r for r in rows if r.get("대상구분") != "child"]

    wb = Workbook()
    wb.remove(wb.active)
    sheet_general(wb, general, title)
    sheet_child(wb, child, title)
    sheet_basis(wb, rows)
    sheet_summary(wb, general, child, title, NOTES)
    wb.save(dst)
    print("생성: {}".format(dst))
    print("  일반 {}종 / 어린이 {}종 / 합계 {}종".format(len(general), len(child), len(rows)))
    print("  정가 합계 {:,}원".format(sum(price(r) for r in rows)))


if __name__ == "__main__":
    main()
